import json
import argparse
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def parse_arguments():
    """
    Parse command line arguments for the JSON to Excel transformer
    """
    parser = argparse.ArgumentParser(
        description="Transform bookawards JSON data into a well-structured Excel file"
    )
    
    parser.add_argument(
        '--input', '-i',
        type=str,
        default='bookawards_result.json',
        help='Path to the input JSON file (default: bookawards_result.json)'
    )
    
    parser.add_argument(
        '--output', '-o',
        type=str,
        default='bookawards.xlsx',
        help='Path to the output Excel file (default: bookawards.xlsx)'
    )
    
    return parser.parse_args()


def load_json_data(json_file):
    """
    Load JSON data from a file
    """
    try:
        with open(json_file, 'r', encoding='utf-8') as file:
            data = json.load(file)
        print(f"Successfully loaded data from {json_file}")
        print(f"Total awards: {len(data)}")
        return data
    except Exception as e:
        print(f"Error loading JSON data: {str(e)}")
        return None


def apply_header_style(sheet, headers):
    """
    Apply styling to the header row
    """
    # Add headers
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def create_awards_overview_sheet(workbook, data):
    """
    Create a sheet with award overview information
    """
    sheet = workbook.create_sheet(title="Awards Overview")
    
    # Define headers
    headers = [
        "Award Name", "Organization", "Registration URL", "Categories",
        "Latest Submission Date", "Enriched Organization", "Enriched Categories", "Enriched Registration URL"
    ]
    
    # Apply styling to header row
    apply_header_style(sheet, headers)
    
    # Add data rows
    row_idx = 2
    for award in data:
        # Basic data
        sheet.cell(row=row_idx, column=1, value=award.get('award_name', ''))
        sheet.cell(row=row_idx, column=2, value=award.get('organization', ''))
        sheet.cell(row=row_idx, column=3, value=award.get('registration_url', ''))
        
        # Make sure categories is a list before joining
        categories = award.get('categories', [])
        if categories and isinstance(categories, list):
            sheet.cell(row=row_idx, column=4, value=', '.join(categories))
        else:
            sheet.cell(row=row_idx, column=4, value='')
        
        # Enriched data (if available)
        if 'enriched_data' in award and isinstance(award['enriched_data'], dict):
            enriched = award['enriched_data']
            sheet.cell(row=row_idx, column=5, value=enriched.get('latestDateOfSubmission', ''))
            sheet.cell(row=row_idx, column=6, value=enriched.get('organization', ''))
            
            # Make sure categories is a list before joining
            enriched_categories = enriched.get('categories', [])
            if enriched_categories and isinstance(enriched_categories, list):
                sheet.cell(row=row_idx, column=7, value=', '.join(enriched_categories))
            else:
                sheet.cell(row=row_idx, column=7, value='')
                
            sheet.cell(row=row_idx, column=8, value=enriched.get('registrationUrl', ''))
        
        row_idx += 1
    
    # Auto-adjust column widths
    for col_idx, _ in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 20
    
    print(f"Created 'Awards Overview' sheet with {row_idx-2} rows")


def create_winning_books_sheet(workbook, data):
    """
    Create a sheet with winning books information
    """
    sheet = workbook.create_sheet(title="Winning Books")
    
    # Define headers
    headers = [
        "Award Name", "Author", "Title", "Publishing Year", "Publisher", "ISBN", "Link"
    ]
    
    # Apply styling to header row
    apply_header_style(sheet, headers)
    
    # Add data rows
    row_idx = 2
    book_count = 0
    
    for award in data:
        award_name = award.get('award_name', '')
        
        if 'enriched_data' in award and isinstance(award['enriched_data'], dict):
            books = award['enriched_data'].get('lastWinningBooks', [])
            
            # Check if books is an iterable
            if books and isinstance(books, list):
                for book in books:
                    if isinstance(book, dict) and book and not all(val == 'Not Available' for val in book.values()):
                        sheet.cell(row=row_idx, column=1, value=award_name)
                        sheet.cell(row=row_idx, column=2, value=book.get('author', ''))
                        sheet.cell(row=row_idx, column=3, value=book.get('title', ''))
                        sheet.cell(row=row_idx, column=4, value=book.get('publishingYear', ''))
                        sheet.cell(row=row_idx, column=5, value=book.get('publisher', ''))
                        sheet.cell(row=row_idx, column=6, value=book.get('isbn', ''))
                        sheet.cell(row=row_idx, column=7, value=book.get('link', ''))
                        
                        row_idx += 1
                        book_count += 1
    
    # If no books were found, add a placeholder row
    if book_count == 0:
        sheet.cell(row=2, column=1, value="No data available")
        print("Warning: No winning books data available, creating empty sheet")
    
    # Auto-adjust column widths
    for col_idx, _ in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 20
    
    print(f"Created 'Winning Books' sheet with {book_count} rows")


def create_competition_sheet(workbook, data):
    """
    Create a sheet with competition information
    """
    sheet = workbook.create_sheet(title="Competition")
    
    # Define headers
    headers = ["Award Name", "Author", "Title"]
    
    # Apply styling to header row
    apply_header_style(sheet, headers)
    
    # Add data rows
    row_idx = 2
    competition_count = 0
    
    for award in data:
        award_name = award.get('award_name', '')
        
        if 'enriched_data' in award and isinstance(award['enriched_data'], dict):
            competitors = award['enriched_data'].get('possibleStrongestCompetitionThisYear', [])
            
            # Ensure competitors is an iterable list
            if competitors and isinstance(competitors, list):
                for competitor in competitors:
                    if isinstance(competitor, dict):
                        sheet.cell(row=row_idx, column=1, value=award_name)
                        sheet.cell(row=row_idx, column=2, value=competitor.get('author', ''))
                        sheet.cell(row=row_idx, column=3, value=competitor.get('title', ''))
                        
                        row_idx += 1
                        competition_count += 1
    
    # If no competition was found, add a placeholder row
    if competition_count == 0:
        sheet.cell(row=2, column=1, value="No data available")
        print("Warning: No competition data available, creating empty sheet")
    
    # Auto-adjust column widths
    for col_idx, _ in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 20
    
    print(f"Created 'Competition' sheet with {competition_count} rows")


def create_categories_sheet(workbook, data):
    """
    Create a sheet with categories information
    """
    sheet = workbook.create_sheet(title="Categories")
    
    # Define headers
    headers = ["Award Name", "Category"]
    
    # Apply styling to header row
    apply_header_style(sheet, headers)
    
    # Add data rows
    row_idx = 2
    category_count = 0
    
    for award in data:
        award_name = award.get('award_name', '')
        categories = award.get('categories', [])
        
        # Ensure categories is an iterable list
        if categories and isinstance(categories, list):
            for category in categories:
                sheet.cell(row=row_idx, column=1, value=award_name)
                sheet.cell(row=row_idx, column=2, value=category)
                
                row_idx += 1
                category_count += 1
    
    # If no categories were found, add a placeholder row
    if category_count == 0:
        sheet.cell(row=2, column=1, value="No data available")
        print("Warning: No category data available, creating empty sheet")
    
    # Auto-adjust column widths
    for col_idx, _ in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 30
    
    print(f"Created 'Categories' sheet with {category_count} rows")


def transform_json_to_excel(json_file, excel_file):
    """
    Transform JSON data to Excel format
    """
    # Load JSON data
    data = load_json_data(json_file)
    if not data:
        return False
    
    try:
        # Create a new workbook
        workbook = Workbook()
        
        # Remove the default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # Create sheets with data
        create_awards_overview_sheet(workbook, data)
        create_winning_books_sheet(workbook, data)
        create_competition_sheet(workbook, data)
        create_categories_sheet(workbook, data)
        
        # Set the first sheet as active
        workbook.active = 0
        
        # Save the workbook
        workbook.save(excel_file)
        
        print(f"\nExcel file '{excel_file}' has been created successfully.")
        return True
    
    except Exception as e:
        print(f"Error creating Excel file: {str(e)}")
        return False


if __name__ == "__main__":
    # Parse command line arguments
    args = parse_arguments()
    
    # Transform JSON to Excel
    if transform_json_to_excel(args.input, args.output):
        print(f"\nTransformation complete.")
        print(f"Output file: {os.path.abspath(args.output)}")
    else:
        print("\nTransformation failed.")
